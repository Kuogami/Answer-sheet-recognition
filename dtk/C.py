import numpy as np
import cv2
import os
import math
from imutils import auto_canny, contours
import openpyxl

# 透视变换要找到变换矩阵。变换矩阵要求：原图的4个点坐标和变换之后的4个点坐标-现在已找到原图的4个点坐标. 需要找到变换之后的4个坐标  先对获取到的4个角点坐标按照一定顺序(顺时针或逆时针)排序
# 为了获取第四个点，编写一个函数seek_points通过两条边相交计算第四个点
def seek_points(A,B,C,D):
    # 计算直线 AB 和 CD 的斜率
    m_ab = (B[1] - A[1]) / (B[0] - A[0]) if B[0] != A[0] else float('inf')
    m_cd = (D[1] - C[1]) / (D[0] - C[0]) if D[0] != C[0] else float('inf')
    # 计算求出第四个点
    x=(m_ab*A[0]-m_cd*C[0]+C[1]-A[1])/(m_ab-m_cd)
    y = m_ab * (x - A[0]) + A[1]
    return [x,y]
# 找到对应的四个点
def order_points(pts):
    rect = np.zeros((6, 2), dtype='float32') # 创建全是0的矩阵, 来接收等下找出来的6个角的坐标
    rect_f = np.zeros((4, 2), dtype='float32')  # 创建全是0的矩阵, 来接收等下找出来的透视变换矩阵的4个角的坐标
    sorted_x = np.argsort(pts[:, 0])
    # 左上、下角的点
    if pts[sorted_x[0]][1] <= pts[sorted_x[1]][1]:
        rect[0]=pts[sorted_x[0]]  # 左上
        rect[1] = pts[sorted_x[1]]  # 左下
    else:
        rect[0] = pts[sorted_x[1]]
        rect[1] = pts[sorted_x[0]]
    # 右上、右中的点
    if pts[sorted_x[5]][1] <= pts[sorted_x[4]][1]:
        rect[2] = pts[sorted_x[5]]  # 右上
        rect[4] = pts[sorted_x[4]]  # 右中
    else:
        rect[2] = pts[sorted_x[4]]
        rect[4] = pts[sorted_x[5]]
    # 使用 np.argsort 获取 y 坐标排序后的索引
    sorted_y = np.argsort(pts[:, 1])
    # 下中的点
    if pts[sorted_y[5]][0] > pts[sorted_y[4]][0]:
        rect[5] = pts[sorted_y[5]]  # 下中
    else:
        rect[5] = pts[sorted_y[4]]  # 下中
    rect[3]=seek_points(rect[1],rect[5],rect[2],rect[4])
    # print(rect[3])
    rect_f[0]=rect[0]
    rect_f[1]=rect[2]
    rect_f[2]=rect[3]
    rect_f[3]=rect[1]
    return rect_f


# 透视变换函数
def four_point_transform(image, pts):
    rect = order_points(pts) # 对输入的4个坐标排序
    # 将排序后的点分别赋值给左上角（tl）、右上角（tr）、右下角（br）和左下角（bl）
    (tl, tr, br, bl) = rect
    widthA = np.sqrt((br[0] - bl[0]) ** 2 + (br[1] - bl[1]) ** 2)# 下两点距离 # 空间中两点的距离
    widthB = np.sqrt((tr[0] - tl[0]) ** 2 + (tr[1] - tl[1]) ** 2)# 上两点距离
    max_width = max(int(widthA), int(widthB))
    heightA = np.sqrt((tr[0] - br[0]) ** 2 + (tr[1] - br[1]) ** 2)# 右两点距离
    heightB = np.sqrt((tl[0] - bl[0]) ** 2 + (tl[1] - bl[1]) ** 2)# 左两点距离
    max_height = max(int(heightA), int(heightB))
    dst = np.array([ # 构造变换之后的对应坐标位置
        [0, 0],
        [max_width - 1, 0],
        [max_width - 1, max_height - 1],
        [0, max_height - 1]], dtype='float32') # 从0开始所以减一
    M = cv2.getPerspectiveTransform(rect, dst) # 计算变换矩阵
    warped = cv2.warpPerspective(image, M, (max_width, max_height)) # 透视变换-图片摆正
    return warped


def main():
    # 设置输出目录
    output_dir = r"D:\dtk"

    # 读取图像
    img = cv2.imread(r"D:\dtk\dtk1.jpg")

    # 将图像转换为灰度图
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    # 对灰度图像应用高斯模糊，减少噪声
    blurred = cv2.GaussianBlur(gray, (7, 7), 0)
    # 保存模糊后的图像
    cv2.imwrite(os.path.join(output_dir, 'blurred.jpg'), blurred)

    # 使用Canny算法进行边缘检测
    edged = cv2.Canny(blurred, 75, 200)
    # 保存边缘检测后的图像
    cv2.imwrite(os.path.join(output_dir, 'edged.jpg'), edged)

    # 检测边缘图像中的轮廓
    cnts = cv2.findContours(edged.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)[0]

    # 创建一个图像副本，用于绘制轮廓
    contours_img = img.copy()
    # 在图像上绘制所有检测到的轮廓
    cv2.drawContours(contours_img, cnts, -1, (0, 0, 255), 3)
    # 保存绘制轮廓后的图像
    cv2.imwrite(os.path.join(output_dir, 'contours_img.jpg'), contours_img)

    # 初始化计数器
    count = 0
    # 确保至少检测到一个轮廓
    if len(cnts) > 0:
        # 根据轮廓面积对轮廓进行降序排序
        cnts = sorted(cnts, key=cv2.contourArea, reverse=True)
        # 根据轮廓周长对轮廓进行降序排序
        cnts = sorted(cnts, key=lambda c: cv2.arcLength(c, True), reverse=True)
        # 遍历排序后的轮廓
        for c in cnts:
            # 计算轮廓的周长
            perimeter = cv2.arcLength(c, True)
            # 对轮廓进行多边形近似
            approx = cv2.approxPolyDP(c, 0.025 * perimeter, True)
            # 如果近似后的轮廓有6个点，认为找到了答题卡的轮廓
            if len(approx) == 6:
                docCnt = approx
                area = cv2.contourArea(c)
                # 如果是第一个找到的轮廓，绘制圆点和编号
                if count == 0:
                    for pt in approx:
                        cv2.circle(img, (int(pt[0][0]), int(pt[0][1])), 15, (0, 0, 255), -1)
                    cv2.putText(img, str(count), (approx[0][0][0] + 10, approx[0][0][1] + 10),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 0, 0), 2)
                # 更新计数器
                count += 1
                # 找到答题卡轮廓后，跳出循环
                break

    # 保存标记了轮廓的图像
    cv2.imwrite(os.path.join(output_dir, 'dd.jpg'), img)

    # 进行透视变换，对图像进行校正
    warped = four_point_transform(gray, docCnt.reshape(6, 2))
    # 保存透视变换后的图像
    cv2.imwrite(os.path.join(output_dir, 'warped.jpg'), warped)

    # 取出涂黑区域
    kernel = np.ones((9, 9), np.uint8)#9 9  图二  9 9
    dilated = cv2.dilate(warped, kernel, iterations=3)#3  图二  1
    cv2.imwrite(os.path.join(output_dir, 'dilated.jpg'), dilated)
    # 腐蚀
    eroded = cv2.erode(dilated, kernel, iterations=3)#3  图二  1
    cv2.imwrite(os.path.join(output_dir, 'eroded.jpg'), eroded)
    ret, marked_area = cv2.threshold(eroded, 173, 255,
                                     cv2.THRESH_BINARY)#173 255 图二  150 255
    cv2.imwrite(os.path.join(output_dir, 'marked_area.jpg'), marked_area)

    # 确定归一化的目标尺寸
    target_width = 100
    target_height = 100

    height, width = warped.shape
    # 计算归一化因子
    x_scale = target_width / width
    y_scale = target_height / height
    # 确定刻度间隔
    scale_x = int(2/x_scale)  # 可以根据图像大小调整
    scale_y = int(2/y_scale) # 可以根据图像大小调整
    # 绘制x轴刻度线和刻度值
    for i in range(0, width, scale_x):
        normalized_x = int(i * x_scale)
        cv2.line(warped, (i, 0), (i, 10), (0, 0, 255), 1)  # 绘制垂直刻度线
        cv2.putText(warped, str(normalized_x), (i, 20), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255), 1)
    # 绘制y轴刻度线和刻度值
    for i in range(0, height, scale_y):
        normalized_y = int(i * y_scale)
        cv2.line(warped, (0, i), (10, i), (0, 0, 255), 1)  # 绘制水平刻度线
        cv2.putText(warped, str(normalized_y), (20, i), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255), 1)

    cv2.imwrite(os.path.join(output_dir, 'warped2.jpg'), warped)

    cnts, h = cv2.findContours(marked_area, cv2.RETR_LIST, cv2.CHAIN_APPROX_NONE)
    # 根据轮廓的x坐标对轮廓进行排序（从左到右）
    cnts = sorted(cnts, key=lambda cnt: cv2.boundingRect(cnt)[0])
    temp1_ans_img=warped
    for i, c in enumerate(cnts):
        cv2.drawContours(temp1_ans_img, cnts, i, (0, 0, 0), 1)
    cv2.imwrite(os.path.join(output_dir, 'temp1_ans_img.jpg'), temp1_ans_img)

    # 定义选项的x坐标和对应的题号
    option_x_coords = {6: 'A', 32: 'A', 58: 'A', 84: 'A',
                       11: 'B', 36: 'B', 62: 'B', 88: 'B',
                       15: 'C', 41: 'C', 67: 'C', 92: 'C',
                       19: 'D', 45: 'D',71: 'D', 97: 'D'}

    # 定义行的y坐标
    row_y_coords = {4.5:1 ,7:2 ,9.5:3 ,12.5:4,15:5,
                    20.5:6 ,23:7 ,25.5:8,28.5:9,31:10,
                    36:11 ,38.5:12 ,41.5:13,44:14,46.5:15,
                    52:16 ,54.5:17 ,57.5:18,59.5:19,62.5:20,
                    67.8:21 ,70.5:22 ,73:23,75.5:24,78.3:25,
                    83.3:26 ,86.2:27 ,89:28,91.5:29,94.2:30
    }

    # 用于存储选项的数组，假设有30行，每行4题
    answers = {i: [] for i in range(1, 106)}

    # 检查每个轮廓
    for cnt in cnts:
        # 获取轮廓的边界框
        x, y, w, h = cv2.boundingRect(cnt)
        tmp_x = 2  # 误差值
        tmp_y = 1.5
        # 计算轮廓的中心点坐标
        center_x = x + w // 2
        center_y = y + h // 2
        center_x *= x_scale
        center_y *= y_scale
        # print(center_x," ",center_y)
        # 检查面积和长宽比
        area = cv2.contourArea(cnt)
        aspect_ratio = float(w) / h if h != 0 else 0
        # print(aspect_ratio,"bus",area)
        if area < 200 or aspect_ratio < 0.2 or aspect_ratio > 5:
            print(area,aspect_ratio)
            continue

        # 确定选项
        for opt_x, opt_label in option_x_coords.items():
            if center_x <= opt_x + tmp_x and center_x + tmp_x >= opt_x:
                option = opt_label
                break
        else:
            continue  # 如果没有匹配的选项，跳过这个轮廓

        # 确定行
        for opt_y, rk in row_y_coords.items():
            if center_y <= opt_y + tmp_y and center_y + tmp_y >= opt_y:
                if center_x>=2 and center_x<=23:
                    ex=0
                elif center_x>=26 and center_x<=48:
                    ex=30
                elif center_x>=54 and center_x<=74:
                    ex=60
                elif center_x>=80 and center_x<=99:
                    ex=90
                row = rk + ex
                break
        else:
            continue  # 如果没有匹配的行，跳过这个轮廓
        answers[row].append(option)
        # 记录答案

    # 创建一个新的Excel工作簿
    wb = openpyxl.Workbook()
    ws = wb.active  # 获取活动工作表

    # 将答案写入工作表
    for question, options in answers.items():  # 题号从1开始
        # 将题号写入A列
        ws[f'A{question}'] = question
        # 将所有选项用逗号分隔，写入B列
        ws[f'B{question}'] = ', '.join(options)
    # 保存工作簿
    output_path = r"D:\dtk\answer.xlsx"  # 输出文件路径
    wb.save(output_path)
    # 打印答案
    for question, options in answers.items():
        print(f"Question {question}: {', '.join(options)}")


if __name__ == "__main__":
    main()